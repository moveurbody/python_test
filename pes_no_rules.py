#!/usr/bin/python
# coding=UTF-8

import json
import logging
import os
import pyodbc
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, colors,Border,Side

import mail

with open('pes_no_rules_cfg.json') as data_file:
    cfg = json.load(data_file)
print cfg["ConnectionStr"]

# Setup log format,level and path
logging.basicConfig(level=logging.DEBUG,
                    filename='debug.log',
                    format='%(asctime)s %(levelname)s %(message)s',
                    datefmt='%y-%m-%d %H:%M:%S')
# Time
localtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
localtime2 = time.strftime('%Y%m%d_%H%M', time.localtime(time.time()))
localtime3 = time.strftime('%Y%m%d', time.localtime(time.time() - 24*60*60))
# Local path
current_exec_path = os.getcwd()
try:
    # connection string
    logging.info("Start to get data")
    logging.debug("Create Connection")
    connStr = pyodbc.connect(str(cfg["ConnectionStr"]))
    cursor = connStr.cursor()

    # Get 6 hours data by tag
    # if count < 0 send mail and provide data
    logging.info("+++ Check status")

    # ErrorFound flag default is false
    ErrorFound = False
    sqlStr = "SELECT [WebLink],[folder],[UID],[serialno],[its_project],[process_time] \
              FROM [PowerExpertSystem].[dbo].[UserTrial_NoRule] \
              WHERE process_time between CONVERT(char(10),GETDATE()-1,120) and CONVERT(char(10),GETDATE(),120) \
              ORDER BY its_project, serialno, process_time"

    cursor.execute(sqlStr)
    rows = cursor.fetchall()
    # If the rows include any data, it means some issue create / reply fail, we need to send alarm
    if rows:
        ErrorFound = True
        logging.debug("Non-defined cases was found.")

        # Inital Excel
        logging.debug("+++ Inital Excel")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Title Format
        title_font = Font(name='Calibri', color=colors.WHITE, size=10)
        title_bg = PatternFill(fill_type='solid', start_color='FF068FFF', end_color='FF068FFF')
        title_border = Border(left=Side(border_style='thin', color='FF000000'),
                              right=Side(border_style='thin', color='FF000000'),
                              top=Side(border_style='thin', color='FF000000'),
                              bottom=Side(border_style='thin', color='FF000000'))
        col = 1
        while col < 11:
            ws.cell(row=1, column=col).font = title_font
            ws.cell(row=1, column=col).fill = title_bg
            ws.cell(row=1, column=col).border = title_border
            col += 1

        # Insert Title
        ws.cell(row=1, column=1).value = "Symptom"
        ws.cell(row=1, column=2).value = "Subsymptom"
        ws.cell(row=1, column=3).value = "Acquire Time"
        ws.cell(row=1, column=4).value = "LogLink"
        ws.cell(row=1, column=5).value = "WebLink"
        ws.cell(row=1, column=6).value = "Folder"
        ws.cell(row=1, column=7).value = "UID"
        ws.cell(row=1, column=8).value = "SerialNo"
        ws.cell(row=1, column=9).value = "ITS Project"
        ws.cell(row=1, column=10).value = "Process Time"

        # Data Format
        data_font = Font(name='Calibri', color=colors.BLACK, size=10)
        data_font_link = Font(name='Calibri', color=colors.BLUE, size=10, underline='single')
        data_border = Border(left=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             top=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'))
        # Insert Data
        row_num = 2
        for row_all in rows:
            # print row_all.WebLink, row_all.folder, row_all.UID, row_all.serialno, row_all.its_project, row_all.process_time
            ws.cell(row=row_num, column=4).value = "=HYPERLINK(E"+str(row_num)+",G"+str(row_num)+")"
            ws.cell(row=row_num, column=5).value = row_all.WebLink
            ws.cell(row=row_num, column=6).value = row_all.folder
            ws.cell(row=row_num, column=7).value = row_all.UID
            ws.cell(row=row_num, column=8).value = row_all.serialno
            ws.cell(row=row_num, column=9).value = row_all.its_project
            ws.cell(row=row_num, column=10).value = row_all.process_time

            col = 1
            while col < 11:
                ws.cell(row=row_num, column=col).font = data_font
                ws.cell(row=row_num, column=col).border = data_border
                col += 1
            col = 4
            while col < 5:
                ws.cell(row=row_num, column=col).font = data_font_link
                col += 1
            row_num += 1

        excel_logpath = 'PES_No_Rules'+localtime2+'.xlsx'
        wb.save(excel_logpath)
        fullpath = current_exec_path+"\\"+excel_logpath
        logging.debug("--- Inital Excel")

        # Send mail
        logging.debug(cfg['mail_receiver'])
        mail.send_mail(cfg['mail_sender'], cfg['mail_receiver'], cfg['mail_subject']+" on "+localtime3, cfg['mail_body'], fullpath)

        # Delete File
        os.remove(fullpath)
    else:
        ErrorFound = False
        logging.debug("No Non-defined Cases")
    logging.info("--- Check status")

except Exception, e:
    print e
    logging.error(e)