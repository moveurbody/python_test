#!/usr/bin/python
# coding=UTF-8

import json
import logging
import os
import pyodbc
import time
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)

import mail

with open('kernelfeedback_issue_fail_cfg.json') as data_file:
    cfg = json.load(data_file)
print cfg["ConnectionStr"]

# Setup log format,level and path
logging.basicConfig(level=logging.DEBUG,
                    filename='debug.log',
                    format='%(asctime)s %(levelname)s %(message)s',
                    datefmt='%y-%m-%d %H:%M:%S')
# Time
localtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
localtime2 = time.strftime('%Y%m%d_%H%M', time.localtime(time.time()))
# Local path
current_exec_path = os.getcwd()
try:
    # connection string
    logging.info("Start to get data")
    logging.debug("Create Connection")
    connStr = pyodbc.connect(str(cfg["ConnectionStr"]))
    cursor = connStr.cursor()

    # Get 6 hours data by tag
    # if count < 0 send mail and provide data
    logging.info("+++ Check status")

    # ErrorFound flag default is false
    ErrorFound = False
    sqlStr = "SELECT LogDetailID, ErrorPatternID, LogReceivedTime, CreatedTime, Tag, ItsProject, LogPath \
              FROM AndFirstException WHERE ErrorPatternID>0 AND LogDetailID IN \
              (SELECT LogDetailID FROM AndFirstExceptionUtd WHERE ItsType IS NULL AND Comments IS NULL) \
              ORDER BY LogDetailID"

    cursor.execute(sqlStr)
    rows = cursor.fetchall()
    # If the rows include any data, it means some issue create / reply fail, we need to send alarm
    if rows:
        ErrorFound = True
        logging.debug("Error Found")

        # Inital Excel
        logging.debug("+++ Inital Excel")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Insert Title
        ws.cell(row=1, column=1).value = "LogDetailID"
        ws.cell(row=1, column=2).value = "ErrorPatternID"
        ws.cell(row=1, column=3).value = "LogReceivedTime"
        ws.cell(row=1, column=4).value = "CreatedTime"
        ws.cell(row=1, column=5).value = "Tag"
        ws.cell(row=1, column=6).value = "ItsProject"
        ws.cell(row=1, column=7).value = "LogPath"
        # Insert Data
        row_num = 2
        for row_all in rows:
            print row_all.LogDetailID, row_all.ErrorPatternID, row_all.LogReceivedTime, row_all.CreatedTime, row_all.Tag, row_all.ItsProject, row_all.LogPath
            ws.cell(row=row_num, column=1).value = row_all.LogDetailID
            ws.cell(row=row_num, column=2).value = row_all.ErrorPatternID
            ws.cell(row=row_num, column=3).value = row_all.LogReceivedTime
            ws.cell(row=row_num, column=4).value = row_all.CreatedTime
            ws.cell(row=row_num, column=5).value = row_all.Tag
            ws.cell(row=row_num, column=6).value = row_all.ItsProject
            ws.cell(row=row_num, column=7).value = row_all.LogPath
            row_num += 1
        excel_logpath = 'KernelFeedBack_'+localtime2+'.xlsx'
        wb.save(excel_logpath)
        fullpath=current_exec_path+"\\"+excel_logpath
        logging.debug("--- Inital Excel")

        # Send mail
        logging.debug(cfg['mail_receiver'])
        mail.send_mail(cfg['mail_sender'], cfg['mail_receiver'], cfg['mail_subject'], cfg['mail_body'], fullpath)

        # Update the time for next run
        sqlStr_update_time = "UPDATE AndFirstException \
                              SET LogReceivedTime= CONVERT(nvarchar(11),GETDATE(),120)+'11:11:11' \
                              WHERE ErrorPatternID>0 \
                              AND LogDetailID IN \
                             (SELECT LogDetailID FROM AndFirstExceptionUtd WHERE  ItsType IS NULL AND Comments IS NULL)\
                             and LogDetailID NOT IN (select LogDetailID from AndItsIssue)"
        logging.debug(sqlStr_update_time)
        logging.debug("Update time for fail cases")
        connStr = pyodbc.connect(str(cfg["ConnectionStr"]))
        cursor = connStr.cursor()
        cursor.execute(sqlStr_update_time)
        connStr.commit()

        # Delete File
        os.remove(fullpath)
    else:
        ErrorFound = False
        logging.debug("No Error Found")
    logging.info("--- Check status")

except Exception, e:
    print e
    logging.error(e)