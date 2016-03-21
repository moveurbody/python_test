#!/usr/bin/python
# coding=UTF-8

import json
import logging
import os
import pyodbc
import time
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)

import mail

with open('log_sync_notification_cfg.json') as data_file:
    cfg = json.load(data_file)
print cfg["ConnectionStr"]

# Setup log format,level and path
logging.basicConfig(level=logging.DEBUG,
                    filename='debug.log',
                    format='%(asctime)s %(levelname)s %(message)s',
                    datefmt='%y-%m-%d %H:%M:%S')
# Time
localtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
localtime2 = time.strftime('%Y%m%d_%H%M', time.localtime(time.time()))
# Local path
current_exec_path = os.getcwd()

# To get time,tag and count in 6 hours data from AndLogDetail
# return the list to rows_logcount
def log_detail_list_count():
    try:
        logging.info("+++ Get all data from AndLogDetail")
        connstr_logcount = pyodbc.connect(str(cfg["ConnectionStr"]))
        cursor_logcount = connstr_logcount.cursor()

        sqlstr_logcount="SELECT Time,\
        MAX(case when tag='HTC_LOG_UPLOAD' then Count else 0 end) as 'HTC_LOG_UPLOAD',\
        MAX(case when tag='HTC_MODEM_RESET' then Count else 0 end) as 'HTC_MODEM_RESET',\
        MAX(case when tag='HTC_PWR_EXPERT' then Count else 0 end) as 'HTC_PWR_EXPERT',\
        MAX(case when tag='LASTKMSG' then Count else 0 end) as 'LASTKMSG'\
        FROM (\
        SELECT CONVERT(char(13),[CreatedTime],120)+':00:00' AS 'Time',tag,COUNT(tag) AS 'Count'\
        FROM [KernelInfo].[dbo].[AndLogDetail]\
        WHERE CreatedTime BETWEEN DATEADD(Hour, -6, GETDATE()) AND GETDATE()\
        GROUP BY CONVERT(char(13),[CreatedTime],120),tag) T\
        GROUP BY Time"
        cursor_logcount.execute(sqlstr_logcount)
        rows_logcount = cursor_logcount.fetchall()
        logging.info("--- Get all data from AndLogDetail")
        return rows_logcount
    except Exception, e:
        print e
        logging.error(e)


try:
    # connection string
    logging.info("Start to get data")
    logging.debug("Create Connection")
    connStr = pyodbc.connect(str(cfg["ConnectionStr"]))
    cursor = connStr.cursor()

    TagList = ('HTC_LOG_UPLOAD', 'HTC_MODEM_RESET', 'HTC_PWR_EXPERT', 'LASTKMSG')

    # Get 6 hours data by tag
    # if count < 0 send mail and provide data
    logging.info("+++ Check status")
    ErrorFound = False
    for tag in TagList:
        sqlStr = "SELECT COUNT(tag) AS 'Count'\
                 FROM [KernelInfo].[dbo].[AndLogDetail]\
                 WHERE tag='"+tag+"' and CreatedTime BETWEEN DATEADD(Hour, -"+str(cfg['CheckHour'])+", GETDATE()) AND GETDATE()\
                 GROUp BY tag"

        cursor.execute(sqlStr)
        rows = cursor.fetchall()
        if not rows:
            ErrorFound = True
        else:
            for row in rows:
                logging.debug(row.Count)
                if row.Count > 0:
                    logging.debug(tag+" count is "+str(row.Count))
                else:
                    ErrorFound = True
                    logging.debug('fail!!!')
                    logging.debug(tag+" count is "+str(row.Count))

    if ErrorFound:
        logging.warning("Fail!!")
        rows_all = log_detail_list_count()
        # Inital Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.cell(row=1, column=1).value = "Time"
        ws.cell(row=1, column=2).value = "HTC_LOG_UPLOAD"
        ws.cell(row=1, column=3).value = "HTC_MODEM_RESET"
        ws.cell(row=1, column=4).value = "HTC_PWR_EXPERT"
        ws.cell(row=1, column=5).value = "LASTKMSG"

        # Insert Data
        row_num = 2
        for row_all in rows_all:
            print row_all.Time, row_all.HTC_LOG_UPLOAD, row_all.HTC_MODEM_RESET, row_all.HTC_PWR_EXPERT, row_all.LASTKMSG
            ws.cell(row=row_num, column=1).value = row_all.Time
            ws.cell(row=row_num, column=2).value = row_all.HTC_LOG_UPLOAD
            ws.cell(row=row_num, column=3).value = row_all.HTC_MODEM_RESET
            ws.cell(row=row_num, column=4).value = row_all.HTC_PWR_EXPERT
            ws.cell(row=row_num, column=5).value = row_all.LASTKMSG
            row_num += 1
            # Create line chart
            c1 = LineChart()
            c1.title = "Log Counts of Tags"
            c1.style = 2
            data = Reference(ws, min_col=2, min_row=1, max_col=5, max_row=8)
            c1.add_data(data, titles_from_data=True)
            ws.add_chart(c1, "A10")
        excel_logpath = 'Detail_'+localtime2+'.xlsx'
        wb.save(excel_logpath)
        fullpath=current_exec_path+"\\"+excel_logpath
        mail.send_mail(cfg['mail_sender'], cfg['mail_receiver'], cfg['mail_subject'], cfg['mail_body'], fullpath)

        # Delete File
        os.remove(fullpath)
    else:
        logging.info("Normal!")
    logging.info("--- Check status")

except Exception, e:
    print e
    logging.error(e)